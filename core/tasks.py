import csv
import io
from datetime import timedelta

from accounts.models import CustomUser
from books.models import Book, BookIssue
from celery import shared_task
from django.db.models import Count
from django.utils import timezone
from schools.models import School


@shared_task
def generate_students_report(school_id, fmt='csv'):
    school = School.objects.get(id=school_id)
    students = CustomUser.objects.filter(school=school, role='student').order_by('grade', 'last_name')
    headers = ['#', 'Username', 'First Name', 'Last Name', 'Grade']
    rows = [[i + 1, s.username, s.first_name, s.last_name, s.grade or ''] for i, s in enumerate(students)]
    return _build_response(f'students_{school.id}', headers, rows, fmt)


@shared_task
def generate_books_report(school_id, fmt='csv'):
    school = School.objects.get(id=school_id)
    books = Book.objects.filter(school=school).select_related('category').order_by('title')
    headers = ['#', 'Title', 'Author', 'Category', 'Total', 'Available', 'Textbook']
    rows = [
        [
            i + 1,
            b.title,
            b.author or '',
            b.category.name if b.category else '',
            b.total_count,
            b.available_count,
            'Yes' if b.is_textbook else 'No',
        ]
        for i, b in enumerate(books)
    ]
    return _build_response(f'books_{school.id}', headers, rows, fmt)


@shared_task
def generate_issues_report(school_id, fmt='csv'):
    school = School.objects.get(id=school_id)
    issues = BookIssue.objects.filter(book__school=school).select_related('book', 'user').order_by('-issued_at')
    headers = ['#', 'Book', 'User', 'Issued At', 'Returned At', 'Status']
    rows = [
        [
            i + 1,
            iss.book.title,
            iss.user.username,
            iss.issued_at,
            iss.returned_at or '',
            'Returned' if iss.is_returned else 'Active',
        ]
        for i, iss in enumerate(issues)
    ]
    return _build_response(f'issues_{school.id}', headers, rows, fmt)


@shared_task
def generate_stats_report(school_id, fmt='csv'):
    school = School.objects.get(id=school_id)
    total_students = CustomUser.objects.filter(school=school, role='student').count()
    total_teachers = CustomUser.objects.filter(school=school, role='teacher').count()
    total_books = Book.objects.filter(school=school).count()
    active_issues = BookIssue.objects.filter(book__school=school, is_returned=False).count()
    returned_issues = BookIssue.objects.filter(book__school=school, is_returned=True).count()
    labels = ['Total Students', 'Total Teachers', 'Total Books', 'Active Loans', 'Returned Books']
    values = [total_students, total_teachers, total_books, active_issues, returned_issues]
    headers = ['Metric', 'Value']
    rows = list(zip(labels, values))
    return _build_response(f'stats_{school.id}', headers, rows, fmt, no_header=True)


@shared_task
def schedule_weekly_news():
    from schools.models import News

    today = timezone.now()
    week_ago = today - timedelta(days=7)
    for school in School.objects.all():
        top_schools = (
            BookIssue.objects.values('book__school__name')
            .filter(issued_at__gte=week_ago)
            .annotate(count=Count('id'))
            .order_by('-count')[:5]
        )
        top_readers = (
            CustomUser.objects.filter(role='student', bookissue__issued_at__gte=week_ago)
            .annotate(count=Count('bookissue'))
            .order_by('-count')[:5]
        )
        title = f'Weekly Active Report — {today.strftime("%Y-%m-%d")}'
        if not News.objects.filter(school=school, title=title, created_at__date=today.date()).exists():
            News.objects.create(
                school=school,
                title=title,
                body='',
                is_published=True,
                template_key='weekly_active',
                template_data={
                    'schools': [{'name': s['book__school__name'], 'count': s['count']} for s in top_schools],
                    'readers': [{'username': r.username, 'grade': r.grade, 'count': r.count} for r in top_readers],
                },
            )


def _build_response(filename, headers, rows, fmt, no_header=False):
    buf = io.BytesIO()
    if fmt == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        wb = Workbook()
        ws = wb.active
        if not no_header:
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            start = 2
        else:
            start = 1
        for r_idx, row in enumerate(rows, start):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3
        wb.save(buf)
        ct = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif fmt == 'pdf':
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        num_cols = len(headers) if headers else 1
        page_size = landscape(A4) if num_cols > 5 else A4
        doc = SimpleDocTemplate(buf, pagesize=page_size, topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()
        elements = [Paragraph('Report', styles['Title']), Spacer(1, 12)]
        table_data = []
        if not no_header:
            table_data.append([str(h) for h in headers])
        for row in rows:
            table_data.append([str(c) if c is not None else '' for c in row])
        if table_data:
            tbl = Table(table_data, repeatRows=1 if not no_header else 0)
            style_cmds = [
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
            if not no_header:
                style_cmds.append(('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A90D9')))
                style_cmds.append(('TEXTCOLOR', (0, 0), (-1, 0), colors.white))
            tbl.setStyle(TableStyle(style_cmds))
            elements.append(tbl)
        doc.build(elements)
        ct = 'application/pdf'
    else:
        writer = csv.writer(buf)
        if not no_header:
            writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        ct = 'text/csv; charset=utf-8-sig'
    return buf.getvalue(), ct

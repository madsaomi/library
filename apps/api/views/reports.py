import csv
import io

from accounts.models import CustomUser
from books.models import Book, BookIssue
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action
from schools.models import School

from api.permissions import IsSchoolAdmin
from api.permissions import IsSuperuser as IsAdmin


class ReportViewSet(viewsets.ViewSet):
    @action(detail=False, methods=['get'], permission_classes=[IsSchoolAdmin])
    def students(self, request):
        fmt = request.query_params.get('output')
        school = request.user.school
        students = CustomUser.objects.filter(school=school, role='student').order_by('grade', 'last_name')
        return self._response(
            request,
            'Students Report',
            fmt,
            ['#', 'Username', 'First Name', 'Last Name', 'Grade'],
            [[i + 1, s.username, s.first_name, s.last_name, s.grade or ''] for i, s in enumerate(students)],
        )

    @action(detail=False, methods=['get'], permission_classes=[IsSchoolAdmin])
    def teachers(self, request):
        fmt = request.query_params.get('output')
        school = request.user.school
        teachers = CustomUser.objects.filter(school=school, role='teacher').order_by('last_name')
        return self._response(
            request,
            'Teachers Report',
            fmt,
            ['#', 'Username', 'First Name', 'Last Name', 'Subject'],
            [[i + 1, t.username, t.first_name, t.last_name, t.subject or ''] for i, t in enumerate(teachers)],
        )

    @action(detail=False, methods=['get'], permission_classes=[IsSchoolAdmin])
    def books(self, request):
        fmt = request.query_params.get('output')
        school = request.user.school
        books = Book.objects.filter(school=school).select_related('category').order_by('title')
        return self._response(
            request,
            'Books Catalog',
            fmt,
            ['#', 'Title', 'Author', 'Category', 'Total', 'Available', 'Textbook'],
            [
                [
                    i + 1,
                    b.title,
                    b.author or '',
                    b.category.name if b.category else '',
                    b.total_count,
                    b.available_count,
                    'Yes' if b.is_textbook else 'No',
                ]
                for i, b in enumerate(books)
            ],
        )

    @action(detail=False, methods=['get'], permission_classes=[IsSchoolAdmin])
    def issues(self, request):
        fmt = request.query_params.get('output')
        school = request.user.school
        issues = BookIssue.objects.filter(book__school=school).select_related('book', 'user').order_by('-issued_at')
        return self._response(
            request,
            'Book Issues Report',
            fmt,
            ['#', 'Book', 'User', 'Issued At', 'Returned At', 'Status'],
            [
                [
                    i + 1,
                    iss.book.title,
                    iss.user.username,
                    iss.issued_at,
                    iss.returned_at or '',
                    'Returned' if iss.is_returned else 'Active',
                ]
                for i, iss in enumerate(issues)
            ],
        )

    @action(detail=False, methods=['get'], permission_classes=[IsSchoolAdmin])
    def stats(self, request):
        fmt = request.query_params.get('output')
        school = request.user.school
        total_students = CustomUser.objects.filter(school=school, role='student').count()
        total_teachers = CustomUser.objects.filter(school=school, role='teacher').count()
        total_books = Book.objects.filter(school=school).count()
        active_issues = BookIssue.objects.filter(book__school=school, is_returned=False).count()
        returned_issues = BookIssue.objects.filter(book__school=school, is_returned=True).count()
        data = [
            ['Metric', 'Value'],
            ['Total Students', total_students],
            ['Total Teachers', total_teachers],
            ['Total Books', total_books],
            ['Active Loans', active_issues],
            ['Returned Books', returned_issues],
        ]
        return self._response(request, f'School Statistics - {school.name}', fmt, data[0], data[1:], no_header=True)

    @action(detail=False, methods=['get'], permission_classes=[IsAdmin])
    def admin_stats(self, request):
        fmt = request.query_params.get('output')
        total_schools = School.objects.count()
        total_students = CustomUser.objects.filter(role='student').count()
        total_teachers = CustomUser.objects.filter(role='teacher').count()
        total_books = Book.objects.count()
        active_loans = BookIssue.objects.filter(is_returned=False).count()
        data = [
            ['Metric', 'Value'],
            ['Total Schools', total_schools],
            ['Total Students', total_students],
            ['Total Teachers', total_teachers],
            ['Total Books', total_books],
            ['Active Loans', active_loans],
        ]
        return self._response(request, 'Global Statistics', fmt, data[0], data[1:], no_header=True)

    def _response(self, request, title, fmt, headers, rows, no_header=False):
        fmt = fmt or 'csv'
        if fmt == 'pdf':
            return self._pdf_response(title, headers, rows, no_header)
        elif fmt == 'xlsx':
            return self._xlsx_response(title, headers, rows, no_header)
        else:
            return self._csv_response(title, headers, rows, no_header)

    def _csv_response(self, title, headers, rows, no_header):
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{title.lower().replace(" ", "_")}.csv"'
        writer = csv.writer(response)
        if not no_header:
            writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return response

    def _xlsx_response(self, title, headers, rows, no_header):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]
        if not no_header:
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            start_row = 2
        else:
            start_row = 1
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{title.lower().replace(" ", "_")}.xlsx"'
        wb.save(response)
        return response

    def _pdf_response(self, title, headers, rows, no_header):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        buf = io.BytesIO()
        num_cols = len(headers) if headers else (len(rows[0]) if rows else 1)
        page_size = landscape(A4) if num_cols > 5 else A4
        doc = SimpleDocTemplate(buf, pagesize=page_size, topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()
        elements = []
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 12))
        table_data = []
        if not no_header:
            table_data.append([str(h) for h in headers])
        for row in rows:
            table_data.append([str(c) if c is not None else '' for c in row])
        if table_data:
            tbl = Table(table_data, repeatRows=1 if not no_header else 0)
            style_cmds = [
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
            if not no_header:
                style_cmds.append(('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A90D9')))
                style_cmds.append(('TEXTCOLOR', (0, 0), (-1, 0), colors.white))
                style_cmds.append(('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'))
            tbl.setStyle(TableStyle(style_cmds))
            elements.append(tbl)
        doc.build(elements)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{title.lower().replace(" ", "_")}.pdf"'
        response.write(buf.getvalue())
        buf.close()
        return response
